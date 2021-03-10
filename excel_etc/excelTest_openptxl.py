#http://openpyxl.readthedocs.io/en/default/index.html

from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active #활성화 되어 있는 시트 선택
#ws = wb.create_sheet(0) #0이면 첫번째 위치에 삽입, 생략시 마지막 위치에 삽입
ws.title="sample"
ws['B2']=42
ws.append([1,2,3])
ws.append([1,2,3])
ws.append([1,2,3])
ws.append([4,4,4])
wb.save("openpyxl1.xlsx")
wb.close()

wb = load_workbook(filename = 'openpyxl1.xlsx')
ws = wb.get_active_sheet()
ws['A1']= 42
ws.cell(row=1, colum=3).value =333
ws.append(['aaa','bbb','ccc'])#새로운 행에 추가됨

print(ws['A1'].value)
print(ws['A2'].value) #저장된 값이 없으면 None으로 출력

ws2 = wb['sample']
print(ws2['A3'].value, ws2['B3'].value, ws2['C3'].value)
print(ws2['A4'].value, ws2['B4'].value, ws2['C4'].value)
print(ws2['A5'].value, ws2['B5'].value, ws2['C5'].value)

wb.save("openpyxl1.xlsx")