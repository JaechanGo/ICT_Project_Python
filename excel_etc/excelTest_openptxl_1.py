from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("diary",0) # 0이면 시트 첫번ㄷ째에 추가, 생략시 마지막 위치에 삽입

data = [('홍길동', 80,70,90),('이기자', 90,60,80),('강기자', 80,80,70)]

r = 1
c = 1

for irum, kor, eng, math in data:
    ws.cell(row=r,column=c).value = irum
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r+=1

ws.cell(row=r, column=1).value = '평'
ws.cell(row=r, column=2).value = '=average(B1:B3)'
ws.cell(row=r, column=3).value = '=average(C1:C3)'
ws.cell(row=r, column=4).value = '=average(D1:D3)'

wb.save("openpyxl2.xlsx")
wb.close()